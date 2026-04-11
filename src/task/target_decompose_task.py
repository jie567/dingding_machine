import calendar
import datetime
import re

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook

from src.task.Task import Task, task_config
from src.util.excel_writer import excel_copy, batch_excel_writer
from src.util.oracle_connect import OracleDataConn


@task_config(
    name="收益任务月初目标分解",
    # ex_time='0 10 * L *',
    task_type="file",
    chat_id="chatdbf714606bde86a565d7bb01767fdd75",
    excel_path='../../resources/excel_template/收益任务分解.xlsx',
    input_path='../../resources/input_excel/航线区域表.xlsx',
    sql_file1='../../resources/sql/月度任务分解.sql',
    sql_file2='../../resources/sql/月度任务分解区域座收.sql'
)
class TargetDecomposeTask(Task):
    def __init__(self):
        super().__init__('收益任务月初目标分解')

    def date_deter(self, cur_date):
        today = datetime.date.today()  if cur_date is None else pd.to_datetime(cur_date)
        next_month = 1 if today.month == 12 else today.month + 1
        next_year = today.year + 1 if today.month == 12 else today.year
        huanqi_month = today.month - 1 if today.month > 1 else 12
        huanqi_year = today.year if today.month > 1 else today.year - 1
        last_year = today.year - 1

        cur_month = datetime.date(today.year, today.month, 1).strftime('%Y-%m-%d')
        cur_month_end = (datetime.date(next_year, next_month, 1) - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        huanqi_month = datetime.date(huanqi_year, huanqi_month, 1).strftime('%Y-%m-%d')
        huanqi_month_end = (datetime.date(today.year, today.month, 1) - datetime.timedelta(days=1)).strftime('%Y-%m-%d')

        tongqi_month = datetime.date(last_year, today.month, 1).strftime('%Y-%m-%d')

        _, last_day = calendar.monthrange(last_year, today.month)
        tongqi_month_end = datetime.date(last_year, today.month, last_day).strftime('%Y-%m-%d')
        return cur_month, cur_month_end, huanqi_month, huanqi_month_end, tongqi_month, tongqi_month_end

    def compute_total_discount(self, input_df: pd.DataFrame, time_discount_map: dict) -> pd.DataFrame:
        df = input_df.copy()
        df['discount_rate'] = df['UPDIS_TIME'].map(time_discount_map)
        if df['discount_rate'].isnull().any():  # 检查是否有时间在时刻系数表中找不到（理论上不应发生，但为了健壮性）
            missing = df.loc[df['discount_rate'].isnull(), 'UPDIS_TIME'].unique()
            print(f"警告：以下时间在时刻系数表中未找到，将用 0 代替: {missing}")
            df['discount_rate'].fillna(0, inplace=True)

        # 计算ELINE  (AB折扣率 * 航班数+BC折扣率 * 航班数)/(AB航班数+BC航班数) 按 ELINE 汇总
        def weighted_mean(group):
            numerator = (group['FLIGHT_NUM'] * group['discount_rate']).sum()
            denominator = group['FLIGHT_NUM'].sum()
            return numerator / denominator if denominator != 0 else 0

        result = df.groupby(['ELINE', 'FLIGHT_NO', 'PLANE'], as_index=False).apply(weighted_mean)
        result.columns = ['ELINE', 'FLIGHT_NO', 'PLANE', 'total_discount']  # 重命名结果列
        return result

    def cal_time_df(self, time_discount_map, cur_time_df, huanqi_time_df):
        db_df3_discount = self.compute_total_discount(cur_time_df, time_discount_map)
        db_df4_discount = self.compute_total_discount(huanqi_time_df, time_discount_map)
        time_df = pd.merge(db_df3_discount, db_df4_discount,
                           on=['ELINE', 'FLIGHT_NO', 'PLANE'], how='left',suffixes=('_cur', '_huanqi'))
        time_df['time_factor'] = (time_df['total_discount_cur'] / time_df['total_discount_huanqi']).round(2)
        # time_df['time_factor'] = (1+(time_df['total_discount_cur']-time_df['total_discount_huanqi']) / time_df['total_discount_huanqi']).round(2)
        time_df.drop(['total_discount_cur', 'total_discount_huanqi'], axis=1, inplace=True)
        return time_df

    def is_English(self, text):
        """判断字符串是否不包含英文字符（返回 True 表示有英文）"""
        if pd.isna(text):
            return False
        return bool(re.search(r'[a-zA-Z]', str(text)))

    def split_segments(self, value):
        """按'/'拆分字符串，返回非空列表"""
        if pd.isna(value) or value == '':
            return []
        return [seg.strip() for seg in str(value).split('/') if seg.strip()]

    def match_sy(self, segment, sy_df,  date_diff):
        """
        在 sy 表中匹配一个子段，返回 (SUM_INCOME, SUM_ASK) 的和
        segment 可能是纯英文（如 "AKU-AAT"）或中英混合（如 "新疆-XUH"）
        """
        if pd.isna(segment):
            return 0.0, 0.0
        parts = segment.split('-', 1)
        if len(parts) != 2:
            return 0.0, 0.0
        chn_part = parts[0].strip()
        eng_part = parts[1].strip()

        if self.is_English(segment) and not self.is_English(chn_part):
            # 中英混合：提取中文省份和英文地点列表 """ is_English 判断字符串是否不包含英文字符（返回 True 表示有英文）"""
            # 过滤 sy_df 满足条件
            mask = (sy_df['UP_PROVINCE'] == chn_part) & \
                   (sy_df['DIS_LOCATION'] == eng_part) & \
                   (sy_df['DATE_DIFF'] == date_diff)
            matched = sy_df.loc[mask]

        elif self.is_English(segment) and not self.is_English(eng_part):
            mask = (sy_df['UP_LOCATION'] == chn_part) & \
                   (sy_df['DIS_PROVINCE'] == eng_part) & \
                   (sy_df['DATE_DIFF'] == date_diff)
            matched = sy_df.loc[mask]
        else:
            # 全英文：直接匹配 SEGMENT
            mask = (sy_df['SEGMENT'] == segment) & \
                   (sy_df['DATE_DIFF'] == date_diff)
            matched = sy_df.loc[mask]

        if matched.empty:
            return 0.0, 0.0
        return matched['SUM_INCOME'].sum(), matched['SUM_ASK'].sum()

    def match_eu(self, segment, eu_df, date_diff):
        """
        在 eu 表中匹配一个子段（纯中文），返回 (SUM_INCOME, SUM_ASK) 的和
        """
        if pd.isna(segment):
            return 0.0, 0.0
        mask = (eu_df['TAG'] == segment) & \
               (eu_df['DATE_DIFF'] == date_diff)
        matched = eu_df.loc[mask]
        if matched.empty:
            return 0.0, 0.0
        return matched['SUM_INCOME'].sum(), matched['SUM_ASK'].sum()

    def process_region(self, region_value, eu_tongqi_df, eu_huanqi_df, sy_tongqi_df, sy_huanqi_df, date_diff):
        """
        处理一个区域字段（可能包含 '/' 分割的多个子段） # 全中文用eu，其他用sy
        返回 (tongqi_income, tongqi_ask, huanqi_income, huanqi_ask)
        """
        segments = self.split_segments(region_value)
        tongqi_income, tongqi_ask = 0.0, 0.0
        huanqi_income, huanqi_ask = 0.0, 0.0

        for seg in segments:
            if not self.is_English(seg):
                # 纯中文 -> 使用 eu 表  且可能含 '/'，但这里 seg 已经是单个子段，直接匹配
                inc, ask = self.match_eu(seg, eu_tongqi_df, date_diff)
                tongqi_income += inc
                tongqi_ask += ask
                inc, ask = self.match_eu(seg, eu_huanqi_df,date_diff)
                huanqi_income += inc
                huanqi_ask += ask
            else:
                # 英文或混合 -> 使用 sy 表
                inc, ask = self.match_sy(seg, sy_tongqi_df, date_diff)
                tongqi_income += inc
                tongqi_ask += ask
                inc, ask = self.match_sy(seg, sy_huanqi_df, date_diff)
                huanqi_income += inc
                huanqi_ask += ask

        return tongqi_income, tongqi_ask, huanqi_income, huanqi_ask

    def add_adjustment_columns(self, df, eu_tongqi_df, eu_huanqi_df, sy_tongqi_df, sy_huanqi_df):
        """
        为主 DataFrame 添加 tongqi 和 huanqi 的 SUM_INCOME 和 SUM_ASK 列
        """
        # 初始化累加列
        df['TONGQI_INCOME_1'] = 0.0
        df['TONGQI_ASK_1'] = 0.0
        df['HUANQI_INCOME_1'] = 0.0
        df['HUANQI_ASK_1'] = 0.0

        df['TONGQI_INCOME_2'] = 0.0
        df['TONGQI_ASK_2'] = 0.0
        df['HUANQI_INCOME_2'] = 0.0
        df['HUANQI_ASK_2'] = 0.0

        # 逐行处理（数据量不大时可接受）
        for idx, row in df.iterrows():
            date_diff = row['DATE_DIFF']
            region1 = row.get('区域1', None)
            region2 = row.get('区域2', None)

            # 处理区域1
            if not pd.isna(region1):
                t_inc, t_ask, h_inc, h_ask = self.process_region(
                    region1, eu_tongqi_df, eu_huanqi_df, sy_tongqi_df, sy_huanqi_df, date_diff
                )
                df.at[idx, 'TONGQI_INCOME_1'] += t_inc
                df.at[idx, 'TONGQI_ASK_1'] += t_ask
                df.at[idx, 'HUANQI_INCOME_1'] += h_inc
                df.at[idx, 'HUANQI_ASK_1'] += h_ask

            # 处理区域2（如果非空）
            if not pd.isna(region2):
                t_inc, t_ask, h_inc, h_ask = self.process_region(
                    region2, eu_tongqi_df, eu_huanqi_df, sy_tongqi_df, sy_huanqi_df, date_diff
                )
                df.at[idx, 'TONGQI_INCOME_2'] += t_inc
                df.at[idx, 'TONGQI_ASK_2'] += t_ask
                df.at[idx, 'HUANQI_INCOME_2'] += h_inc
                df.at[idx, 'HUANQI_ASK_2'] += h_ask
        # 区域1
        df['RASK_1_TONGQI'] = df['TONGQI_INCOME_1'] / df['TONGQI_ASK_1'].replace(0, np.nan)
        df['RASK_1_HUANQI'] = df['HUANQI_INCOME_1'] / df['HUANQI_ASK_1'].replace(0, np.nan)
        # 区域2
        df['RASK_2_TONGQI'] = df['TONGQI_INCOME_2'] / df['TONGQI_ASK_2'].replace(0, np.nan)
        df['RASK_2_HUANQI'] = df['HUANQI_INCOME_2'] / df['HUANQI_ASK_2'].replace(0, np.nan)

        # 计算调整系数
        df['region_factor'] = np.where(
            (df['RASK_1_TONGQI'].notna()) & (df['RASK_1_HUANQI'].notna()),
            df['RASK_1_TONGQI'] / df['RASK_1_HUANQI'],
            np.nan
        )
        # 区域2不为空 区域系数重新计算
        mask_region2 = df['区域2'].notna() & df['RASK_2_TONGQI'].notna() & df['RASK_2_HUANQI'].notna()
        df.loc[mask_region2, 'region_factor'] = (0.5 * df.loc[mask_region2, 'region_factor']
                                                 + 0.5 * (df.loc[mask_region2, 'RASK_2_TONGQI'] / df.loc[mask_region2, 'RASK_2_HUANQI']))

        df.drop(['TONGQI_INCOME_1', 'TONGQI_ASK_1', 'HUANQI_INCOME_1', 'HUANQI_ASK_1', 'RASK_1_TONGQI', 'RASK_1_HUANQI',
                 'TONGQI_INCOME_2', 'TONGQI_ASK_2', 'HUANQI_INCOME_2', 'HUANQI_ASK_2', 'RASK_2_TONGQI', 'RASK_2_HUANQI'], axis=1, inplace=True)
        return df

    def execute_task(self, conn, cur_date=None, cur_month_end=None, huanqi_month=None, huanqi_month_end=None,
                     tongqi_month=None, tongqi_month_end=None, data_need_insert = True):
        task_config = getattr(self.__class__, '_task_config', {})
        output_excel_path = task_config.get('excel_path', '../../resources/excel_template/收益任务分解.xlsx')
        input_file = task_config.get('input_path', '../../resources/input_excel/航线区域表.xlsx')
        sql_file1 = task_config.get('sql_file1', '../../resources/sql/月度任务分解.sql')
        sql_file2 = task_config.get('sql_file2', '../../resources/sql/月度任务分解区域座收.sql')
        # 以测算2025年12月为例  环期 huanqi_month为2025.11.01  huanqi_month_end为11.30
        # tongqi_month =  2024.12.01  tongqi_month_end =  2024.12.31  tongqi_huangqi_month_end = 2024.12.31
        # 日期计算
        if huanqi_month is None and tongqi_month is None:
            cur_month, cur_month_end, huanqi_month, huanqi_month_end, tongqi_month, tongqi_month_end = self.date_deter(cur_date)

        print(" cur_month, cur_month_end: ",cur_month, cur_month_end, '\n',
              "huanqi_month,huanqi_month_end:",huanqi_month,huanqi_month_end, '\n',
              "tongqi_month,tongqi_month_end:",tongqi_month,tongqi_month_end)
        with open(sql_file1, 'r', encoding='utf-8') as f:
            sql_content = f.read()
        sql_commands = [cmd.strip() for cmd in sql_content.split(';') if cmd.strip()]

        replacements = {
            'huanqi_month_end': huanqi_month_end,
            'huanqi_month': huanqi_month,
            'tongqi_month_end': tongqi_month_end,
            'tongqi_month': tongqi_month,
            'cur_month_end': cur_month_end,
            'cur_month': cur_month
        }
        df_dict = {}
        for idx, sql_command in enumerate(sql_commands, start=1):
            for placeholder, value in replacements.items():
                pattern = r'\b' + re.escape(placeholder) + r'\b'
                sql_command = re.sub(pattern, value, sql_command)
            df_dict[idx] = conn.query_as_df(sql_command)

        # 区域调整系数计算 huanqi_detail_df 环期分天明细座收    huangqi_rask_df 环期实际月座收  cur_time_df  当期环期时刻班数  huanqi_time_df 环期时刻班数
        huanqi_detail_df = df_dict[1][~df_dict[1]['FLIGHT_NO'].astype(str).str.contains('[a-zA-Z]', na=False)]
        execl_area_df = pd.read_excel(input_file, sheet_name='航线明细', usecols=[0, 1, 2])
        huanqi_detail_df = pd.merge(huanqi_detail_df, execl_area_df, how='left', on='ELINE')

        # 全中文用eu，其他用sy
        # 区域季节调整系数 = 同期区域座收（同期本月）2412 /  2511 环期区域座收（当期上月）
        with open(sql_file2, 'r', encoding='utf-8') as f:
            sql_content = f.read()
        eu_sql, sy_sql = sql_content.split(';')
        eu_tongqi_sql = eu_sql.replace('start_date', tongqi_month).replace('end_date', tongqi_month_end)
        eu_huanqi_sql = eu_sql.replace('start_date', huanqi_month).replace('end_date', huanqi_month_end)
        eu_tongqi_df = conn.query_as_df(eu_tongqi_sql)
        eu_huanqi_df = conn.query_as_df(eu_huanqi_sql)

        sy_tongqi_sql = sy_sql.replace('start_date', tongqi_month).replace('end_date', tongqi_month_end)
        sy_huanqi_sql = sy_sql.replace('start_date', huanqi_month).replace('end_date', huanqi_month_end)
        sy_tongqi_df = conn.query_as_df(sy_tongqi_sql)
        sy_huanqi_df = conn.query_as_df(sy_huanqi_sql)

        huanqi_result_df = self.add_adjustment_columns(huanqi_detail_df, eu_tongqi_df, eu_huanqi_df, sy_tongqi_df, sy_huanqi_df)
        print('聚合之前 huanqi_result_df.isna().sum()：\n', huanqi_result_df['region_factor'].isna().sum())
        huanqi_result_df = huanqi_result_df.groupby(['ELINE', 'FLIGHT_NO', 'PLANE'], as_index=False)['region_factor'].mean()
        print('聚合之后 huanqi_result_df.isna().sum()：\n', huanqi_result_df.isna().sum())


        # 时刻调整系数计算
        cur_time_df = df_dict[3][~df_dict[3]['FLIGHT_NO'].astype(str).str.contains('[a-zA-Z]', na=False)]
        huanqi_time_df = df_dict[4][~df_dict[4]['FLIGHT_NO'].astype(str).str.contains('[a-zA-Z]', na=False)]
        excel_time_df = pd.read_excel(input_file, sheet_name='时刻系数', dtype={'时间': str})
        excel_time_df['时间'] = excel_time_df['时间'].astype(str).str.rsplit(':', n=1).str[0]
        time_discount_map = dict(zip(excel_time_df['时间'], excel_time_df['折扣']))

        time_df = self.cal_time_df(time_discount_map, cur_time_df, huanqi_time_df)
        print('len(time_df): ', len(time_df), ',time_df.isna().sum(): ', time_df.isna().sum())

        cur_time_group = cur_time_df.groupby(['ELINE', 'FLIGHT_NO', 'PLANE', 'SEGMENT_TYPE']).agg({
            'FLIGHT_NUM': 'sum',
            'AVG_DIST': 'mean'
        })
        cur_time_group = cur_time_group.groupby(['ELINE', 'FLIGHT_NO', 'PLANE']).agg({
            'FLIGHT_NUM': 'mean',
            'AVG_DIST': 'mean'
        })
        time_df = pd.merge(time_df, cur_time_group, on=['ELINE', 'FLIGHT_NO', 'PLANE'], how='left')

        huangqi_rask_df = df_dict[2][~df_dict[2]['FLIGHT_NO'].astype(str).str.contains('[a-zA-Z]', na=False)]
        merged_df = pd.merge(huangqi_rask_df, huanqi_result_df, how='left', on=['ELINE', 'FLIGHT_NO', 'PLANE'])
        merged_df = pd.merge(time_df, merged_df, how='left', on=['ELINE', 'FLIGHT_NO', 'PLANE'])
        # 首月无计划座收采用下面的
        # merged_df['HUANQI_PLAN_RASK'] = np.nan
        # merged_df['PLANED_RASK'] =  merged_df['time_factor'] * merged_df['HUANQI_ACT_RASK'] * merged_df['region_factor']
        print('len(merged_df): ', len(merged_df), ',merged_df.isna().sum(): ', merged_df.isna().sum())

        # 结果拼接输出
        huangqi_planed_rask_sql = f'''
        select
        ELINE, FLIGHT_NO, PLANE, MONTH_PLANED_RASK as HUANQI_PLAN_RASK
        from AIR_SJ_JY.dwd_revenue_plan_task_mi
        where dep_month = '{huanqi_month[5:7].lstrip('0')}'
        '''
        huangqi_planed_rask_df = conn.query_as_df(huangqi_planed_rask_sql)
        merged_df = pd.merge(merged_df, huangqi_planed_rask_df, how='left' ,on= ['ELINE', 'FLIGHT_NO', 'PLANE'])

        cols_to_fill = ['region_factor', 'time_factor', 'HUANQI_PLAN_RASK', 'HUANQI_ACT_RASK']
        merged_df.fillna({col: 0 for col in cols_to_fill}, inplace=True)

        merged_df['PLANED_RASK'] = (
                merged_df['HUANQI_ACT_RASK']
                .clip(
                    lower=merged_df['HUANQI_PLAN_RASK'] * 0.8,
                    upper=merged_df['HUANQI_PLAN_RASK'] * 1.1
                )
                * merged_df['region_factor']
                * merged_df['time_factor']
        )
        print('最终结果len(merged_df): ', len(merged_df), ',merged_df.isna().sum(): ', merged_df.isna().sum())
        lower_limit = 0.5
        upper_limit = 2

        mask = (
                (merged_df['time_factor'] < lower_limit) | (merged_df['time_factor'] > upper_limit) |
                (merged_df['region_factor'] < lower_limit) | (merged_df['region_factor'] > upper_limit)
        )

        invalid_rows = merged_df[mask]
        if not invalid_rows.empty:
            print("超出范围 (0.5 - 2) 的行：")
            print(invalid_rows)


        new_columns_order = ['ELINE', 'FLIGHT_NO', 'PLANE', 'HUANQI_ACT_RASK', 'HUANQI_PLAN_RASK', 'region_factor',
                             'time_factor', 'PLANED_RASK', 'FLIGHT_NUM', 'AVG_DIST']
        merged_df = merged_df[new_columns_order]
        new_excel_path = excel_copy(output_excel_path, current_date=cur_month[0:7])
        results = []
        results.append((merged_df, 'Sheet1'))
        batch_excel_writer(new_excel_path, results, special_sheets='明细表')
        wb = load_workbook(new_excel_path)
        ws = wb['Sheet1']
        cell = ws.cell(row=1, column=1)
        cell.value = f'''{cur_month[0:4]}年{cur_month[5:7].replace('0', '')}月任务分解表'''
        wb.save(new_excel_path)
        self.file_list.append(new_excel_path)

        if data_need_insert: #默认插入
            self.insert(merged_df, conn, cur_month)

    def insert(self, merged_df, conn, cur_month):
        new_columns_order = ['ELINE', 'FLIGHT_NO', 'PLANE', 'dep_month', 'region_factor', 'time_factor', 'PLANED_RASK']
        insert_df = merged_df[['ELINE', 'FLIGHT_NO', 'PLANE', 'region_factor', 'time_factor', 'PLANED_RASK']].copy()
        insert_df['dep_month'] = cur_month[5:7].lstrip('0')
        insert_df = insert_df[new_columns_order]
        print("插入数据样例：")
        print(insert_df.head())
        # print("\n数据类型：")
        # print(insert_df.dtypes)
        if insert_df.isna().any().any():
            raise ValueError('insert_df中不可以有NaN值，merge into不成功')
        insert_sql = '''
MERGE INTO AIR_SJ_JY.dwd_revenue_plan_task_mi TGT
        USING (
        SELECT 
        :1 AS ELINE,
        :2 AS flight_no,
        :3 AS PLANE,
        :4 AS DEP_MONTH,
        :5 AS region_factor,
        :6 AS time_factor,
        :7 AS MONTH_PLANED_RASK
         FROM DUAL
        ) SRC 
        ON (
        TGT.ELINE       = SRC.ELINE
        AND TGT.flight_no   = SRC.flight_no
        AND TGT.PLANE       = SRC.PLANE
        AND TGT.DEP_MONTH   = SRC.DEP_MONTH)
WHEN MATCHED THEN
		UPDATE SET 
        TGT.region_factor      = SRC.region_factor,
        TGT.time_factor        = SRC.time_factor,
        TGT.MONTH_PLANED_RASK  = SRC.MONTH_PLANED_RASK
WHEN NOT MATCHED THEN
    INSERT (
        ELINE, 
        flight_no, 
        PLANE, 
        DEP_MONTH, 
        region_factor, 
        time_factor, 
        MONTH_PLANED_RASK)
    VALUES (
        SRC.ELINE, 
        SRC.flight_no, 
        SRC.PLANE, 
        SRC.DEP_MONTH, 
        SRC.region_factor, 
        SRC.time_factor, 
        SRC.MONTH_PLANED_RASK)
        '''
        conn.batch_insert(insert_sql, insert_df)


if __name__ == '__main__':
    oracle_conn = OracleDataConn()
    t = TargetDecomposeTask()
    # t.execute(oracle_conn, cur_date='2025-11-01', cur_month_end='2025-11-30')
    t.execute(oracle_conn, cur_date='2025-12-01', cur_month_end='2025-12-31')
    # t.execute(oracle_conn, cur_date='2026-01-01', cur_month_end='2026-01-31')
    # t.execute(oracle_conn, cur_date='2026-02-01', cur_month_end='2026-02-28')
