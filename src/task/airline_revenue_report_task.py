import datetime
import pandas as pd
from src.task import sql_airline_detail_str, airline_sql_str
from src.task.Task import Task, task_config

from openpyxl import load_workbook
from openpyxl.styles import Alignment
import chinese_calendar as calendar

from src.util.excel_writer import excel_copy, batch_excel_writer
from src.util.oracle_connect import OracleDataConn


@task_config(
    name="航线效益日报",
    ex_time="0 16 * * *",
    task_type="file",
    chat_id="chatdbf714606bde86a565d7bb01767fdd75",
    excel_path ='./resources/excel_template/航线效益日报.xlsx'
)
class AirlineRevenueReportTask(Task):
    def __init__(self):
        super().__init__('航线效益日报')

    def execute_once(self, conn, excel_path, yesterday_str):
        sql_str = sql_airline_detail_str(yesterday_str)
        df = conn.query_as_df(sql_str)
        existing_columns = ['JHWF_ELINECN2', 'HXXZ', 'JHWF_FLIGHT_NO', "PLANE1", "FLT", 'JH_FT_TIME', "FT_TIME",
                            "DISASK", "BKASK", "KPSR_ATI", "FUEL_ATI",
                            "SUBSIDY_ATI", 'SUBSIDY', "HYSR_ATI", "HYBCSR_ATI", "YZSR_ATI", "TPSR_ATI", "YSSR_ATI",
                            "INCOME_TOTAL",
                            "VCOST", "FYJE110", "FYJE5", "FYJE18", "FYJE26", "FYJE20", "FYJE21", "FYJE15", "FYJE9",
                            "FYJE17", "FYJE113",
                            "FYJE127", "FYJE142", "FYJE163", "FYJE162", "FYJE487", "KPQZSY", "SUBSIDY_ATI",
                            "XJ_SUBSIDY", "BSP_FEE", "JSXT_FEE",
                            "ZF_FEE", "AGENT_FEE", 'XINJIANG_BAODI', "BRANCH_SUBSIDY", 'AFTERTAX_INCOME_TOTAL',
                            'PRETAX_INCOME_TOTAL']
        df_detail = df[existing_columns]

        sql_str = airline_sql_str(yesterday_str)
        df = conn.query_as_df(sql_str)
        if df.empty:
            self.error = f"【Error】表Tb_market_finance_detail 没有 {yesterday_str} 日期的数据"
            raise Exception(self.error)
        grouped = df.groupby('PLANE2')
        c909_df = grouped.get_group('C909')
        c909_df = c909_df.drop(['PLANE2', 'PLANE1'], axis=1)
        c909_df = c909_df.sort_values(by='含税小时收入', ascending=False)
        n_cols = c909_df.shape[1]
        c909_end_row = (['小计'] + [c909_df.iloc[:, i].mean() if i not in [0, n_cols - 1] else '' for i in
                                    range(1, n_cols - 8)]
                        + [c909_df.iloc[:, i].sum() if i not in [0, n_cols - 1] else '' for i in
                           range(n_cols - 8, n_cols - 1)] + [''])

        c909_df = pd.concat([c909_df, pd.DataFrame([c909_end_row], columns=c909_df.columns)], ignore_index=True)
        c909_df['边贡率'] = c909_df['边贡率'].apply(lambda x: f"{x:.0%}" if pd.notnull(x) else x)
        c909_df['客座率'] = c909_df['客座率'].apply(lambda x: f"{x:.0%}" if pd.notnull(x) else x)

        c909_df['边贡率'] = c909_df['边贡率'].astype('str')
        c909_df['客座率'] = c909_df['客座率'].astype('str')

        kongke_df = grouped.get_group('空客')
        kongke_df = kongke_df.drop(['PLANE2'], axis=1)
        n_cols = kongke_df.shape[1]
        kongke_df = kongke_df.sort_values(by=['PLANE1', '含税小时收入'], ascending=False)
        kongke_end_row = (
                ['小计'] + [''] + [kongke_df.iloc[:, i].mean() if i not in [0, n_cols - 1] else '' for i in
                                   range(2, n_cols - 8)]
                + [kongke_df.iloc[:, i].sum() if i not in [0, n_cols - 1] else '' for i in
                   range(n_cols - 8, n_cols - 1)] + [''])
        kongke_df = pd.concat([kongke_df, pd.DataFrame([kongke_end_row], columns=kongke_df.columns)],
                              ignore_index=True)
        kongke_df['边贡率'] = kongke_df['边贡率'].apply(lambda x: f"{x:.0%}" if pd.notnull(x) else x)
        kongke_df['客座率'] = kongke_df['客座率'].apply(lambda x: f"{x:.0%}" if pd.notnull(x) else x)

        kongke_df['边贡率'] = kongke_df['边贡率'].astype('str')
        kongke_df['客座率'] = kongke_df['客座率'].astype('str')

        zh_cn_columns = ['航线', '航线性质', '航班对', '机型', '班次', '编排小时', '实飞小时', '分摊座公里',
                         '客公里',
                         '税后客票收入', '税后燃油收入', '税后航补收入',
                         '疆内支线补贴', '税后货邮自营收入', '税后货邮包仓收入', '税后逾重收入', '预估税后退票收入',
                         '运输收入', '总收入', '变动成本', '航油消耗',
                         '起降费', '头等舱服务费', '餐食和机供品', '民航发展基金', '航班延误费', '外站驻组费',
                         '客货邮行赔偿费', '系统使用费', '代理手续费', '飞行小时费',
                         '飞机及发动机维修成本', '主营业务税金及附加', '客舱服务费', '机组调度费', '客票签转损益',
                         '税后航线补贴收入', '疆内补贴', 'BSP数据处理费', '结算系统使用费',
                         '支付手续费', '销售代理费', '新疆保底补贴', '支线补贴', '不含税总收入', '含税总收入']
        detail_special_cols = ['航线', '航线性质', '航班对', '机型', '班次', 'BSP数据处理费', '结算系统使用费',
                               '支付手续费', '销售代理费']
        df_detail.columns = zh_cn_columns
        df_detail.loc[df_detail['航班对'] == '2867/2868', '变动成本'] = 215210
        for col in df_detail.columns:
            if col not in detail_special_cols:
                df_detail[col] = df_detail[col].round(2)

        part_special_cols = ['JHWF_ELINECN2', 'PLANE1', '边贡率', '客座率', 'JHWF_FLIGHT_NO']

        for col in c909_df.columns:
            if col in ['含税座收', '不含税座收']:
                c909_df[col] = c909_df[col].round(3)
                kongke_df[col] = kongke_df[col].round(3)
            elif col not in part_special_cols:
                c909_df[col] = c909_df[col].round(2)
                kongke_df[col] = kongke_df[col].round(2)
        results = []
        results.append((df_detail, '明细表'))
        results.append((c909_df, 'ARJ'))
        results.append((kongke_df, '空客'))
        new_excel_path = excel_copy(excel_path, current_date=yesterday_str)
        batch_excel_writer(new_excel_path, results, special_sheets='明细表')
        sheet_names = ['空客', 'ARJ']

        wb = load_workbook(new_excel_path)
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws = wb['明细表']
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
        wb.save(new_excel_path)

        return new_excel_path

    def execute_task(self, conn, today = None):
        task_config = getattr(self.__class__, '_task_config', {})
        excel_path = task_config.get('excel_path')

        if today is not None and type(today) is str:
            new_excel_path = self.execute_once(conn, excel_path, today)
            self.file_list.append(new_excel_path)
        elif today is None:
            today = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
            yesterday = today - datetime.timedelta(days=1)
            yesterday_str = yesterday.strftime('%Y-%m-%d')
            if calendar.is_workday(today) and not calendar.is_workday(yesterday):
                while (not calendar.is_workday(yesterday)):
                    new_excel_path = self.execute_once(conn, excel_path, yesterday_str)
                    self.file_list.append(new_excel_path)
                    yesterday = yesterday - datetime.timedelta(days=1)
                    yesterday_str = yesterday.strftime('%Y-%m-%d')
                if calendar.is_workday(yesterday):
                    new_excel_path = self.execute_once(conn, excel_path, yesterday_str)
                    self.file_list.append(new_excel_path)
            elif calendar.is_workday(today):
                new_excel_path = self.execute_once(conn, excel_path, yesterday_str)
                self.file_list.append(new_excel_path)
        else:
            raise  Exception("today must be str or None")




if __name__ == '__main__':
    oracle_conn = OracleDataConn()
    airline_task = AirlineRevenueReportTask()
    print(airline_task.ex_time)
    print(airline_task.excel_path)

    # airline_task.execute(oracle_conn)
    # airline_task.execute(oracle_conn, today ='2026-02-01')