import datetime

import pandas as pd

from src.task import uatp_sql_str
from src.task.Task import Task, task_config
from src.util.excel_writer import excel_copy
from src.util.oracle_connect import OracleDataConn
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from src.util.util_function import cell_to_indices

@task_config(
    name="uatp周任务",
    ex_time="30 9 * * 4",
    task_type="file",
    chat_id="chatdbf714606bde86a565d7bb01767fdd75",
    excel_path='./resources/excel_template/UATP销售数据需求字段.xlsx',
    sql_file='./resources/sql/UATP_detail.sql'
)
class uatp_week_task(Task):
    def __init__(self):
        super().__init__('uatp周任务')
        self.input_excel_path = 'W:\03-财务市场共享数据\01 日报基础数据\截图基础数据维护250720.xlsx'
    def uatp_sum_job_run(self, conn):
        print(f"{datetime.datetime.now()}开始进行uatp_sum_job_run任务")
        try:
            startdate_str = 'sysdate-7'
            enddate_str = 'sysdate-1'
            this_week_sql_str = uatp_sql_str(startdate_str, enddate_str)

            startdate_str = 'sysdate-14'
            enddate_str = 'sysdate-8'
            last_week_sql_str = uatp_sql_str(startdate_str, enddate_str)

            df1 = conn.query_as_df(this_week_sql_str)
            df2 = conn.query_as_df(last_week_sql_str)
            merged_df = pd.merge(df1, df2, on='GROUPBY_KEY', how='outer', suffixes=('_this_week', '_last_week'))

            custom_order = ['管总', '西部', '东部', '北部', '新疆']
            merged_df['GROUPBY_KEY'] = merged_df['GROUPBY_KEY'].astype(str)
            merged_df['sort_key'] = merged_df['GROUPBY_KEY'].apply(
                lambda x: custom_order.index(x) if x in custom_order else len(custom_order)
            )
            merged_df = merged_df.sort_values(by=['sort_key', 'GROUPBY_KEY']).drop(columns='sort_key')

            print("【info】 uatp_sum_job_run sql查询结束")
            return merged_df
        except Exception as e:
            raise Exception("【ERROR】 uatp_sum_job_run sql:" + e)

    def uatp_year_job_run(self, conn):
        print(f"{datetime.datetime.now()}开始进行uatp_year_job_run任务")
        try:
            startdate_str = "TRUNC(SYSDATE, 'YYYY')"
            enddate_str = 'sysdate-1'
            year_sql_str = uatp_sql_str(startdate_str, enddate_str)
            df = conn.query_as_df(year_sql_str)

            custom_order = ['管总', '西部', '东部', '北部', '新疆']
            df['GROUPBY_KEY'] = df['GROUPBY_KEY'].astype(str)
            df['sort_key'] = df['GROUPBY_KEY'].apply(
                lambda x: custom_order.index(x) if x in custom_order else len(custom_order)
            )
            df = df.sort_values(by=['sort_key', 'GROUPBY_KEY']).drop(columns='sort_key')
            filter_cols = ['GROUPBY_KEY', 'COUNT(A.PNAME)', 'SUM_SEG_FARE', 'AVG_SEG_FARE', 'RRPK', 'DUFEI_RATE',
                           'RUOJINGZHENG_RATE', 'QIANGJINGZHENG_RATE']
            filter_df = df[filter_cols]
            print("【info】 uatp_year_job_run sql查询结束")
            return filter_df
        except Exception as e:
            raise Exception("【ERROR】 uatp_year_job_run sql:" + e)

    def execute_task(self, conn):
        task_config = getattr(self.__class__, '_task_config', {})
        excel_path = task_config.get('excel_path', '../../resources/excel_template/UATP销售数据需求字段.xlsx')
        sql_file = task_config.get('sql_file', '../../resources/sql/UATP_detail.sql')

        sum_df = self.uatp_sum_job_run(conn)
        print("【info】uatp_sum_job_run:", sum_df)
        year_df = self.uatp_year_job_run(conn)
        print("【info】uatp_year_job_run:", year_df)
        year_df.drop(columns=year_df.columns[0], axis=1, inplace=True)

        with open(sql_file, 'r', encoding='utf-8') as f:
            sql_content = f.read()
        db_df = conn.query_as_df(sql_content)

        new_excel_path = excel_copy(excel_path)
        sum_start_cell = 'B4'
        year_start_cell = 'AE4'
        start_row, start_col = cell_to_indices(sum_start_cell)
        sum_sheet_name, detail_sheet_name = '数据明细', '周报'

        book = load_workbook(new_excel_path)
        ws = book[sum_sheet_name]
        for r_idx, row in enumerate(dataframe_to_rows(sum_df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
        start_row, start_col = cell_to_indices(year_start_cell)
        for r_idx, row in enumerate(dataframe_to_rows(year_df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        if detail_sheet_name in book.sheetnames:
            ws = book[detail_sheet_name]
            ws.delete_rows(2, ws.max_row - 1)
            print(f"【info】{detail_sheet_name}－sheet清理完成")
        else:
            ws = book.create_sheet(detail_sheet_name)
            print(f"{detail_sheet_name}不存在，已创建")
            ws.append([None] * len(db_df.columns))

        for r_idx, row in enumerate(dataframe_to_rows(db_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        book.save(new_excel_path)
        self.file_list.append(new_excel_path)

if __name__ == '__main__':
    oracle_conn = OracleDataConn()
    u_task = uatp_week_task()
    u_task.execute(oracle_conn)