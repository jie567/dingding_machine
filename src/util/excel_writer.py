import datetime

import pandas as pd
import xlrd
import xlwt
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import shutil,os
from xlutils.copy import copy
from pathlib import Path
from openpyxl.utils import get_column_letter

def excel_copy(excel_path, current_date = None):
    if current_date is None:
        current_date = datetime.datetime.now().strftime('%Y%m%d')
    base_name, ext = os.path.splitext(os.path.basename(excel_path))
    base_name = base_name.split("-")[0]
    base_name = base_name.split("_")[0]
    new_excel_path = os.path.join(Path(__file__).parent.parent.parent, "resources", "excels", f"{base_name}-{current_date}{ext}")
    os.makedirs(os.path.dirname(new_excel_path), exist_ok=True)

    if os.path.exists(excel_path):
        shutil.copy2(excel_path, new_excel_path)
        print(f"已复制模板文件到: {new_excel_path}")
    else:
        wb = Workbook()
        wb.save(new_excel_path)
        print(f"创建新工作簿: {new_excel_path}")
    return new_excel_path


def write_to_excel_col_fastest(excel_path: str, sheet_name: str, excel_df: pd.DataFrame,
                               start_row: int = 3, start_col: int = 1, with_head=False) -> None:
    """
    高性能写入Excel（openpyxl直接单元格赋值）
    :param excel_path:   Excel文件路径
    :param sheet_name:   工作表名
    :param excel_df:     要写入的DataFrame
    :param start_row:    起始行号（从1开始）
    :param start_col:    起始列号（从1开始）
    :param with_head:     是否写入表头（True时表头占据start_row行，数据从start_row+1行开始）
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # 1. 写入表头（如果需要）
    if with_head:
        for j, col_name in enumerate(excel_df.columns):
            cell_addr = f"{get_column_letter(start_col + j)}{start_row}"
            ws[cell_addr] = col_name
        data_start_row = start_row + 1   # 数据起始行下移一行
    else:
        data_start_row = start_row

    # 2. 写入数据（使用numpy数组，避免DataFrame索引开销）
    values = excel_df.values
    for i in range(values.shape[0]):
        for j in range(values.shape[1]):
            val = values[i, j]
            if pd.notna(val):   # 跳过NaN/None，保持单元格干净
                cell_addr = f"{get_column_letter(start_col + j)}{data_start_row + i}"
                ws[cell_addr] = val

    wb.save(excel_path)
    print(f"成功写入 {excel_path}，行范围 {data_start_row}~{data_start_row + values.shape[0] - 1}，列范围 {start_col}~{start_col + values.shape[1] - 1}")

def batch_excel_writer(excel_path, results, special_sheets = 'Sheet2', hidden_sheets = None):

    """
    批量写入所有结果到Excel
    :param excel_path: Excel文件路径
    :param results: 包含(df, sheet_name)的元组列表
    """
    hidden_sheets = ["走势源数据", '提前无票']
    book = load_workbook(excel_path)

    for df, sheet_name in results:
        print(f"【INFO】正在写入的sheet: {sheet_name}")
        is_special = sheet_name in special_sheets
        start_row = 2 if is_special else 3

        if sheet_name in book.sheetnames:
            ws = book[sheet_name]
        else:
            ws = book.create_sheet(sheet_name)
            ws.append([None] * len(df.columns))
            print(f"{sheet_name}不存在，已创建")
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

        print(f"数据从第{start_row}行开始写入")
        print(f"{datetime.datetime.now()}:成功写入: {sheet_name} ({len(df)}行数据)")

        if hidden_sheets is not None and sheet_name in hidden_sheets:
            ws.sheet_state = 'hidden'
            print(f"{excel_path}-{sheet_name}已经隐藏")
    book.save(excel_path)
    print(f"{datetime.datetime.now()}:所有结果已保存至: {excel_path}")
    return excel_path

def batch_excel_writer2(excel_path, results, special_sheets='Sheet2', hidden_sheets=None):
    MAX_EXCEL_ROWS = 1048570

    if hidden_sheets is None:
        hidden_sheets = ["走势源数据", '提前无票']

    book = load_workbook(excel_path)

    for df, original_sheet_name in results:
        print(f"【INFO】正在处理的sheet: {original_sheet_name}, 数据行数: {len(df)}")

        # 检查是否需要拆分
        if len(df) <= MAX_EXCEL_ROWS:
            write_single_sheet(book, df, original_sheet_name, special_sheets, hidden_sheets)
        else:
            # 需要拆分数据
            num_chunks = (len(df) // MAX_EXCEL_ROWS) + 1
            print(f"数据超过Excel最大行数，将拆分为 {num_chunks} 个sheet")

            for i in range(num_chunks):
                start_idx = i * MAX_EXCEL_ROWS
                end_idx = min((i + 1) * MAX_EXCEL_ROWS, len(df))

                # 获取数据块
                chunk_df = df.iloc[start_idx:end_idx]

                # 生成sheet名称
                if i == 0:
                    chunk_sheet_name = original_sheet_name
                else:
                    chunk_sheet_name = f"{original_sheet_name}_{i + 1}"

                print(f"写入分块: {chunk_sheet_name}, 行数: {len(chunk_df)}")

                # 写入单个sheet
                write_single_sheet(book, chunk_df, chunk_sheet_name, special_sheets, hidden_sheets)

    book.save(excel_path)
    print(f"{datetime.datetime.now()}:所有结果已保存至: {excel_path}")
    return excel_path


def batch_excel_writer_xls(excel_path, results, special_sheets='Sheet2'):
    """
    批量写入所有结果到 .xls 格式 Excel (使用 xlwt)
    :param excel_path: Excel文件路径 (.xls 格式)
    :param results: 包含(df, sheet_name)的元组列表
    :param special_sheets: 特殊sheet名称，数据从第2行开始写入(其他从第3行)
    """
    # 检查文件是否存在
    try:
        rb = xlrd.open_workbook(excel_path, formatting_info=True)
        wb = copy(rb)
        existing_sheets = rb.sheet_names()
    except FileNotFoundError:
        wb = xlwt.Workbook()
        existing_sheets = []

    current_sheets = existing_sheets[:]

    for df, sheet_name in results:
        print(f"【INFO】正在写入的sheet: {sheet_name}")
        is_special = sheet_name in special_sheets
        start_row = 2 if is_special else 3

        if sheet_name in current_sheets:
            ws = wb.get_sheet(current_sheets.index(sheet_name))
        else:
            ws = wb.add_sheet(sheet_name)
            current_sheets.append(sheet_name)
            # 模拟原函数中 append([None] * len(df.columns)) 的行为： 在第一行写入空值，保证 start_row > 1 时前面有空白行
            for col in range(len(df.columns)):
                ws.write(0, col, '')
            print(f"{sheet_name}不存在，已创建")

        data = df.values
        for r_idx, row in enumerate(data):
            excel_row = start_row + r_idx  # 实际写入的 1-based 行号
            for c_idx, value in enumerate(row):
                ws.write(excel_row - 1, c_idx, value)  # xlwt 使用 0-based 索引

        print(f"数据从第{start_row}行开始写入")
        print(f"{datetime.datetime.now()}:成功写入: {sheet_name} ({len(df)}行数据)")

    wb.save(excel_path)
    print(f"{datetime.datetime.now()}:所有结果已保存至: {excel_path}")
    return excel_path

def write_single_sheet(book, df, sheet_name, special_sheets, hidden_sheets):
    print(sheet_name)
    """写入单个sheet的辅助函数"""
    is_special = sheet_name in special_sheets if isinstance(special_sheets,
                                                            (list, tuple)) else sheet_name == special_sheets
    start_row = 2 if is_special else 3

    if sheet_name in book.sheetnames:
        ws = book[sheet_name]
    else:
        ws = book.create_sheet(sheet_name)
        ws.append(list(df.columns))

    print('sheet加载完成，开始写入')
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value

    print(f"数据从第{start_row}行开始写入")
    print(f"{datetime.datetime.now()}:成功写入: {sheet_name} ({len(df)}行数据)")

    # 处理隐藏sheet
    hidden_list = hidden_sheets if isinstance(hidden_sheets, (list, tuple)) else [hidden_sheets]
    if sheet_name in hidden_list:
        ws.sheet_state = 'hidden'
        print(f"{sheet_name}已经隐藏")

def excel_sheet_hidden(excel_path, sheet_list):
    try:
        book = load_workbook(excel_path)
        for sheet_name in sheet_list:
            if sheet_name in book.sheetnames:
                ws = book[sheet_name]
                ws.sheet_state = 'hidden'
            print(f"{excel_path}：{sheet_name}已经隐藏")
    except Exception as e:
        print(e)
    finally:
        book.save(excel_path)


if __name__ == '__main__':
    # hist_excel_path = '../../resources/excels/历史评价-20250815.xlsm'
    # hidden_sheet_list = ['走势源数据', '提前无票']
    # excel_sheet_hidden(hist_excel_path, hidden_sheet_list)
    print(Path(__file__).parent.parent)
